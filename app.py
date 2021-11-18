from tkinter import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
import os


def main_screen():
    global screen
    screen = Tk()
    screen.geometry("1000x600")
    Label(text="").pack()
    Button(text="login", command=login).pack()
    Button(text="register", command=register).pack()
    Button(text="Exit", command=delete_all).pack()
    screen.mainloop()


def register():
    global screen_register
    screen_register = Toplevel(screen)
    screen_register.title = "Register    "
    screen_register.geometry("500x500")
    global user_name
    global password
    user_name = StringVar()
    password = StringVar()
    Label(screen_register, text="Name").pack()
    global user_name_entry
    user_name_entry = Entry(screen_register, textvariable=user_name)
    user_name_entry.pack()
    Label(screen_register, text="Password").pack()
    global password_entry
    password_entry = Entry(screen_register, textvariable=password)
    password_entry.pack()
    Button(screen_register, text="Register", command=register_user).pack()


def register_user():
    userNameInfo = user_name.get()
    passwordInfo = password.get()
    file = open(userNameInfo, "w")
    file.write(userNameInfo)
    file.write("\n")
    file.write(passwordInfo)
    file.close()
    user_name_entry.delete(0, END)
    password_entry.delete(0, END)
    Label(screen_register, text="Registeration Success", fg="green").pack()
    Button(screen_register,text="Leave",command=delete_register_screen).pack()
def delete_register_screen():
    screen_register.destroy()



def login():
    global screen_login
    screen_login = Toplevel(screen)
    screen_login.title("Log In")
    screen_login.geometry("500x500")
    global user_name_verify
    global password_verify
    user_name_verify = StringVar()
    password_verify = StringVar()

    Label(screen_login, text="Pleasse Enter your name*").pack()
    global user_name_entry1
    user_name_entry1 = Entry(screen_login, textvariable=user_name_verify)
    user_name_entry1.pack()

    Label(screen_login, text="Please Enter your password  *").pack()
    global password_entry1
    password_entry1 = Entry(screen_login, textvariable=password_verify)
    password_entry1.pack()
    Button(screen_login, text="Login ", command=login_verify).pack()


def login_verify():
    user_name_to_verify = user_name_verify.get()
    password_to_verify = password_verify.get()
    password_entry1.delete(0, END)
    user_name_entry1.delete(0, END)
    list_of_files = os.listdir()
    if user_name_to_verify in list_of_files:
        file_to_verify = open(user_name_to_verify, "r")
        verify = file_to_verify.read().splitlines()
        file_to_verify.close()
        if password_to_verify == verify[1]:
            login_success()

        else:
            password_not_recognized()
    else:
        user_not_found()


def login_success():
    screen_login.destroy()
    global login_success_screen
    login_success_screen = Toplevel(screen)
    login_success_screen.geometry("1000x1000")
   # Creating the workbook as new
    create = Workbook("project.xlsx")
    create.save("project.xlsx")
    # loading the workbook
    global ws
    global wb
    global name
    global loan_amount
    global loan_year
    global annual_rate
    wb = load_workbook("project.xlsx")
    ws = wb.active
    ws.merge_cells("D1:F1")
    ws["D1"] = "Loan Managment System"

    ws.append(["Name ", "Loan Amount", "Loan Year", "Number Of Payment", "Annual Rate",
              "Monthly Payment", "Annual Payment", "Total Cost", "Total interest rate"])
    maxRow = ws.max_row
    maxColumn = ws.max_column
    print(maxRow, maxColumn)

    Label(login_success_screen, text="LOAN MANAGMENT SYSTEM",
          width=45).grid(row=0, column=1, sticky=E)

    Label(login_success_screen, text="Enter Your Name",
          width=45).grid(row=1, column=0, sticky=E)
    name = Entry(login_success_screen)
    name.grid(row=1, column=1, sticky=E)

    Label(login_success_screen, text="Enter Your Loan Amount($$$)",
          width=45).grid(row=2, column=0, sticky=E)
    loan_amount = Entry(login_success_screen)
    loan_amount.grid(row=2, column=1, sticky=E)

    Label(login_success_screen, text="Enter Your Loan Year(Date)",
          width=45).grid(row=3, column=0, sticky=E)
    loan_year = Entry(login_success_screen)
    loan_year.grid(row=3, column=1, sticky=E)

    Label(login_success_screen, text="Enter the Annual Rate(Percent)",
          width=45).grid(row=4, column=0, sticky=E)
    annual_rate = Entry(login_success_screen)
    annual_rate.grid(row=4, column=1, sticky=E)
    # Separator
    # Label(text="  as",width=450,height=20 ).grid(row=4,column=0,sticky=E)

    Label(login_success_screen, text="Outputs", width=45).grid(row=6, column=1)
    Button(login_success_screen, text="Save Data", width=34,
           bg="purple", command=save).grid(row=11, column=1)
    Button(login_success_screen, text="Exit", width=34,
           bg="purple", command=exit_login_succes).grid(row=12, column=1)
def exit_login_succes():
    login_success_screen.destroy()


def save():
    maxRow = ws.max_row+1
    ws['A'+str(maxRow)] = name.get()
    ws['B'+str(maxRow)] = loan_amount.get()
    ws['C'+str(maxRow)] = loan_year.get()
    ws['D'+str(maxRow)] = int(str(loan_year.get()))*12
    ws['E'+str(maxRow)] = f"{annual_rate.get()}%"
    ws['F'+str(maxRow)] = f"=-PMT(E{maxRow}/12,D{maxRow},B{maxRow})"
    ws['G'+str(maxRow)] = f"=F{maxRow}*12"
    ws['H'+str(maxRow)] = f"=D{maxRow}*F{maxRow}"
    ws['I'+str(maxRow)] = f"=H{maxRow}-B{maxRow}"

    wb.save("project.xlsx")

    Label(login_success_screen, text="Loan Amount\t\t" +
          ws[f'B{maxRow}'].value, width=45).grid(row=6, column=0, sticky=E)
    # Formula for monthly payment
    annualInterestRate = int(str(annual_rate.get()))/100
    monthlyInterestRate = annualInterestRate/12
    numberOfPayment = int(str(loan_year.get()))*12
    totalAmountOfPayment = int(str(loan_amount.get()))
    formulaOne = (1+monthlyInterestRate)**numberOfPayment
    top = totalAmountOfPayment*monthlyInterestRate*formulaOne
    bottom = formulaOne-1
    monthlyPayment = top/bottom
    monthlyPayment = round(monthlyPayment, 5)
    Label(login_success_screen, text="Monthly Payment\t\t" +
          str(monthlyPayment)+"$", width=45).grid(row=6, column=2, sticky=E)

    Label(login_success_screen, text="Number Of Payment\t\t" +
          str(numberOfPayment)+"", width=45).grid(row=6, column=1, sticky=E)

    annualPayment = monthlyPayment*12
    annualPayment = round(annualPayment, 4)
    Label(login_success_screen, text="Annual Payment\t\t" +
          str(annualPayment)+"$", width=45).grid(row=10, column=0, sticky=E)
    totalCost = numberOfPayment * monthlyPayment
    totalCost = round(totalCost, 4)
    Label(login_success_screen, text="Total Cost\t\t" +
          str(totalCost)+"$", width=45).grid(row=10, column=1, sticky=E)
    totalInterest = totalCost-totalAmountOfPayment
    totalInterest = round(totalInterest, 4)
    Label(login_success_screen, text="Total Interest " +
          str(totalInterest), width=45).grid(row=10, column=2, sticky=E)

    print("Saved the data")


def password_not_recognized():
    global password_not_recognized_screen
    password_not_recognized_screen = Toplevel(screen_login)
    password_not_recognized_screen.title("Success")
    password_not_recognized_screen.geometry("200x200")
    Label(password_not_recognized_screen,
          text="Password Not Recognized  ").pack()
    Button(password_not_recognized_screen, text="Ok", command=delete4).pack()


def user_not_found():

    global user_not_found_screen
    user_not_found_screen = Toplevel(screen_login)
    user_not_found_screen.title("Success")
    user_not_found_screen.geometry("200x200")
    Label(user_not_found_screen, text="User Not FOund  ").pack()
    Button(user_not_found_screen, text="Ok", command=delete5).pack()


def delete3():
    login_success_screen.destroy()


def delete4():
    password_not_recognized_screen.destroy()


def delete5():
    user_not_found_screen.destroy()

def delete_all():
    screen.destroy()
main_screen()